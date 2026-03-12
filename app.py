from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from io import BytesIO
import zipfile

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

APP_TITLE = 'ERS Timesheet'

ACTIVITY_COLUMNS = [
    'Selecionar', 'Cliente', 'Projeto', 'ID Atividade', 'Atividade', 'Recurso', 'Início', 'Fim'
]
WEEKLY_COLUMNS = [
    'ID Atividade', 'Projeto', 'Atividade', 'Recurso', 'Dom', 'Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'EPT'
]
DETAIL_COLUMNS = [
    'ID Lançamento', 'Data', 'Projeto', 'Atividade', 'Recurso', 'Comentário', 'Restrito', 'EPT', 'Horas Reconhecidas'
]
REPORT_COLUMNS = [
    'Profissional', 'Data', 'Projeto', 'Atividade', 'Recurso', 'Horas', 'Horas Reconhecidas', 'Comentário', 'Restrito', 'EPT'
]
DAY_MAP = [('Dom', 6), ('Seg', 0), ('Ter', 1), ('Qua', 2), ('Qui', 3), ('Sex', 4), ('Sáb', 5)]
HEADER_FILL = PatternFill('solid', fgColor='EDEDED')
GREEN_FILL = PatternFill('solid', fgColor='E6F4EA')
YELLOW_FILL = PatternFill('solid', fgColor='FFF8D9')


def example_activities() -> pd.DataFrame:
    return pd.DataFrame([
        [True, 'PANCO', '2019 PANCO FRENTE TM AMS', 229, 'AMS Chamados', 'HILDA DE OLIVEIRA CONDE', date(2026, 3, 1), date(2026, 3, 31)],
        [True, 'PANCO', '2019 PANCO FRENTE TM AMS', 100, 'Melhoria 56 - Automatização da ordem do frete', 'HILDA DE OLIVEIRA CONDE', date(2026, 3, 1), date(2026, 3, 31)],
        [True, 'CORPORATE', '1019 CORPORATE FRENTE TM AMS', 96, 'AMS Chamados Janeiro 2024', 'Consultor', date(2026, 3, 1), date(2026, 3, 31)],
        [False, 'UNIPAR', '2026 UNIPAR TM', 117, 'Testes Integrados', 'HILDA DE OLIVEIRA CONDE', date(2026, 3, 1), date(2026, 3, 31)],
    ], columns=ACTIVITY_COLUMNS)


def blank_activities() -> pd.DataFrame:
    return pd.DataFrame([[False, '', '', '', '', '', None, None]], columns=ACTIVITY_COLUMNS)


def blank_weekly() -> pd.DataFrame:
    return pd.DataFrame([], columns=WEEKLY_COLUMNS)


def blank_details() -> pd.DataFrame:
    return pd.DataFrame([], columns=DETAIL_COLUMNS)


def first_day_of_week(ref: date) -> date:
    # semana começa no domingo para ficar mais próxima do modelo enviado
    return ref - timedelta(days=(ref.weekday() + 1) % 7)


def ensure_state() -> None:
    if 'activities_df' not in st.session_state:
        st.session_state.activities_df = blank_activities()
    if 'weekly_df' not in st.session_state:
        st.session_state.weekly_df = blank_weekly()
    if 'details_df' not in st.session_state:
        st.session_state.details_df = blank_details()
    if 'week_start' not in st.session_state:
        st.session_state.week_start = first_day_of_week(date.today())
    if 'professional' not in st.session_state:
        st.session_state.professional = ''
    if 'manager' not in st.session_state:
        st.session_state.manager = ''
    if 'simple_project' not in st.session_state:
        st.session_state.simple_project = ''
    if 'simple_activity' not in st.session_state:
        st.session_state.simple_activity = ''
    if 'status' not in st.session_state:
        st.session_state.status = 'Em aberto'


def normalize_date(value):
    if value in ('', None) or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def normalize_df(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ''
    return out[columns]


def sync_weekly_from_activities() -> None:
    activities = normalize_df(st.session_state.activities_df, ACTIVITY_COLUMNS).copy()
    activities = activities[activities['Selecionar'] == True].copy()  # noqa: E712
    rows = []
    for _, row in activities.iterrows():
        rows.append({
            'ID Atividade': row['ID Atividade'],
            'Projeto': row['Projeto'],
            'Atividade': row['Atividade'],
            'Recurso': row['Recurso'],
            'Dom': 0.0,
            'Seg': 0.0,
            'Ter': 0.0,
            'Qua': 0.0,
            'Qui': 0.0,
            'Sex': 0.0,
            'Sáb': 0.0,
            'EPT': '',
        })
    old = normalize_df(st.session_state.weekly_df, WEEKLY_COLUMNS)
    if not old.empty:
        for new_row in rows:
            matches = old[
                (old['ID Atividade'].astype(str) == str(new_row['ID Atividade']))
                & (old['Projeto'].astype(str) == str(new_row['Projeto']))
                & (old['Atividade'].astype(str) == str(new_row['Atividade']))
            ]
            if not matches.empty:
                for col in ['Dom', 'Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'EPT']:
                    new_row[col] = matches.iloc[-1][col]
    st.session_state.weekly_df = pd.DataFrame(rows, columns=WEEKLY_COLUMNS)
    sync_details_from_weekly()


def build_entries_from_weekly(weekly_df: pd.DataFrame, week_start: date, status: str) -> pd.DataFrame:
    records = []
    launch_id = 1
    weekly_df = normalize_df(weekly_df, WEEKLY_COLUMNS)
    for _, row in weekly_df.iterrows():
        for label, offset in DAY_MAP:
            hours = pd.to_numeric(row.get(label, 0), errors='coerce')
            hours = 0.0 if pd.isna(hours) else float(hours)
            if hours > 0:
                records.append({
                    'ID Lançamento': launch_id,
                    'Data': week_start + timedelta(days=offset),
                    'Projeto': str(row.get('Projeto', '')),
                    'ID Atividade': str(row.get('ID Atividade', '')),
                    'Atividade': str(row.get('Atividade', '')),
                    'Recurso': str(row.get('Recurso', '')),
                    'Horas': hours,
                    'Situação': status,
                    'EPT': str(row.get('EPT', '')),
                })
                launch_id += 1
    return pd.DataFrame(records)


def sync_details_from_weekly() -> None:
    entries = build_entries_from_weekly(st.session_state.weekly_df, st.session_state.week_start, st.session_state.status)
    old = normalize_df(st.session_state.details_df, DETAIL_COLUMNS)
    detail_rows = []
    for _, row in entries.iterrows():
        match = old[(old['Projeto'].astype(str) == str(row['Projeto']))
                    & (old['Atividade'].astype(str) == str(row['Atividade']))
                    & (pd.to_datetime(old['Data'], errors='coerce').dt.date == row['Data'])]
        if not match.empty:
            existing = match.iloc[-1].to_dict()
            existing['ID Lançamento'] = row['ID Lançamento']
            detail_rows.append(existing)
        else:
            detail_rows.append({
                'ID Lançamento': row['ID Lançamento'],
                'Data': row['Data'],
                'Projeto': row['Projeto'],
                'Atividade': row['Atividade'],
                'Recurso': row['Recurso'],
                'Comentário': '',
                'Restrito': 'Não',
                'EPT': row['EPT'],
                'Horas Reconhecidas': row['Horas'],
            })
    st.session_state.details_df = pd.DataFrame(detail_rows, columns=DETAIL_COLUMNS)


def weekly_metrics(weekly_df: pd.DataFrame) -> dict[str, float]:
    weekly_df = normalize_df(weekly_df, WEEKLY_COLUMNS)
    result = {}
    for label, _ in DAY_MAP:
        result[label] = float(pd.to_numeric(weekly_df.get(label, pd.Series(dtype=float)), errors='coerce').fillna(0).sum())
    result['Total'] = sum(result.values())
    return result


def validate_weekly(weekly_df: pd.DataFrame) -> list[str]:
    weekly_df = normalize_df(weekly_df, WEEKLY_COLUMNS)
    warnings = []
    for label, _ in DAY_MAP:
        total = float(pd.to_numeric(weekly_df.get(label, pd.Series(dtype=float)), errors='coerce').fillna(0).sum())
        if total > 24:
            warnings.append(f'O dia {label} está com {total:.2f} horas. Isso passa de 24h.')
        elif total > 12:
            warnings.append(f'O dia {label} está com {total:.2f} horas. Vale revisar.')
    return warnings


def build_reports() -> tuple[pd.DataFrame, pd.DataFrame]:
    entries = build_entries_from_weekly(st.session_state.weekly_df, st.session_state.week_start, st.session_state.status)
    details = normalize_df(st.session_state.details_df, DETAIL_COLUMNS).copy()
    if entries.empty:
        empty = pd.DataFrame(columns=REPORT_COLUMNS)
        return empty, empty

    if not details.empty:
        details['Data'] = details['Data'].apply(normalize_date)
    merged = entries.merge(details[['ID Lançamento', 'Comentário', 'Restrito', 'Horas Reconhecidas']], on='ID Lançamento', how='left')
    merged['Horas Reconhecidas'] = pd.to_numeric(merged['Horas Reconhecidas'], errors='coerce').fillna(merged['Horas'])
    merged['Comentário'] = merged['Comentário'].fillna('')
    merged['Restrito'] = merged['Restrito'].fillna('Não')
    merged['EPT'] = merged['EPT'].fillna('')
    merged['Profissional'] = st.session_state.professional

    report = merged[['Profissional', 'Data', 'Projeto', 'Atividade', 'Recurso', 'Horas', 'Horas Reconhecidas', 'Comentário', 'Restrito', 'EPT']].copy()
    report = report.sort_values(['Data', 'Projeto', 'Atividade']).reset_index(drop=True)

    simple = report.copy()
    if st.session_state.simple_project:
        simple = simple[simple['Projeto'] == st.session_state.simple_project].copy()
    if st.session_state.simple_activity:
        simple = simple[simple['Atividade'] == st.session_state.simple_activity].copy()

    return simple.reset_index(drop=True), report


def workbook_from_report(df: pd.DataFrame, title: str, week_start: date, manager: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'
    ws.sheet_view.showGridLines = False
    widths = [22, 14, 30, 32, 24, 10, 18, 38, 12, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A3'] = 'Profissional:'
    ws['B3'] = st.session_state.professional
    ws['D3'] = 'Gestor:'
    ws['E3'] = manager
    ws['A4'] = 'Semana:'
    ws['B4'] = f"{week_start.strftime('%d/%m/%Y')} a {(week_start + timedelta(days=6)).strftime('%d/%m/%Y')}"
    ws['D4'] = 'Status:'
    ws['E4'] = st.session_state.status

    for cell in ['A3', 'B3', 'D3', 'E3', 'A4', 'B4', 'D4', 'E4']:
        ws[cell].font = Font(size=11)

    for c, header in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=6, column=c, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left')

    if df.empty:
        ws['A7'] = 'Nenhum lançamento encontrado.'
    else:
        for r, row in enumerate(df.itertuples(index=False), start=7):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(r, c, value)
                if c == 2 and value:
                    cell.number_format = 'dd/mm/yyyy'
                if c in (6, 7):
                    cell.number_format = '0.00'

    footer_row = max(len(df) + 8, 9)
    ws.cell(footer_row, 1, 'Total de horas')
    total_hours = float(pd.to_numeric(df.get('Horas', pd.Series(dtype=float)), errors='coerce').fillna(0).sum())
    ws.cell(footer_row, 6, total_hours)
    ws.cell(footer_row, 6).number_format = '0.00'
    ws.cell(footer_row, 1).font = Font(bold=True)
    ws.cell(footer_row, 6).font = Font(bold=True)
    ws.cell(footer_row, 1).fill = GREEN_FILL
    ws.cell(footer_row, 6).fill = GREEN_FILL

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def save_json_payload() -> bytes:
    payload = {
        'professional': st.session_state.professional,
        'manager': st.session_state.manager,
        'week_start': st.session_state.week_start.isoformat(),
        'status': st.session_state.status,
        'simple_project': st.session_state.simple_project,
        'simple_activity': st.session_state.simple_activity,
        'activities': normalize_df(st.session_state.activities_df, ACTIVITY_COLUMNS).astype(object).where(pd.notnull(st.session_state.activities_df), None).to_dict(orient='records'),
        'weekly': normalize_df(st.session_state.weekly_df, WEEKLY_COLUMNS).astype(object).where(pd.notnull(st.session_state.weekly_df), None).to_dict(orient='records'),
        'details': normalize_df(st.session_state.details_df, DETAIL_COLUMNS).astype(object).where(pd.notnull(st.session_state.details_df), None).to_dict(orient='records'),
    }
    return json.dumps(payload, ensure_ascii=False, default=str, indent=2).encode('utf-8')


def load_json_payload(file_bytes: bytes) -> None:
    data = json.loads(file_bytes.decode('utf-8'))
    st.session_state.professional = data.get('professional', '')
    st.session_state.manager = data.get('manager', '')
    st.session_state.week_start = datetime.fromisoformat(data.get('week_start')).date()
    st.session_state.status = data.get('status', 'Em aberto')
    st.session_state.simple_project = data.get('simple_project', '')
    st.session_state.simple_activity = data.get('simple_activity', '')
    st.session_state.activities_df = pd.DataFrame(data.get('activities', []), columns=ACTIVITY_COLUMNS)
    st.session_state.weekly_df = pd.DataFrame(data.get('weekly', []), columns=WEEKLY_COLUMNS)
    details = pd.DataFrame(data.get('details', []), columns=DETAIL_COLUMNS)
    if not details.empty:
        details['Data'] = pd.to_datetime(details['Data'], errors='coerce').dt.date
    st.session_state.details_df = details


st.set_page_config(page_title=APP_TITLE, layout='wide')
ensure_state()

st.title(APP_TITLE)
st.caption('Aplicativo local em Streamlit para a ERS: 3 telas, controle semanal e geração dos relatórios simples e compostos.')

with st.sidebar:
    st.header('Configurações')
    st.text_input('Profissional', key='professional', placeholder='Ex.: Andréia Conde')
    st.text_input('Gestor / aprovador', key='manager', placeholder='Ex.: Manager ERS')
    st.date_input('Início da semana', key='week_start', format='DD/MM/YYYY')
    st.selectbox('Status', ['Em aberto', 'Enviado para aprovação', 'Aprovado', 'Rejeitado'], key='status')

    col_a, col_b = st.columns(2)
    with col_a:
        if st.button('Carregar exemplo', use_container_width=True):
            st.session_state.activities_df = example_activities()
            st.session_state.professional = 'HILDA DE OLIVEIRA CONDE'
            st.session_state.manager = 'Gestor ERS'
            st.session_state.week_start = date(2026, 3, 1)
            st.session_state.status = 'Em aberto'
            sync_weekly_from_activities()
            if not st.session_state.weekly_df.empty:
                st.session_state.weekly_df.loc[0, ['Seg']] = [8.0]
                st.session_state.weekly_df.loc[1, ['Seg', 'Qui']] = [3.0, 4.0]
                st.session_state.weekly_df.loc[2, ['Seg', 'Ter']] = [2.0, 2.0]
                sync_details_from_weekly()
                if not st.session_state.details_df.empty:
                    st.session_state.details_df.loc[:, 'Comentário'] = [
                        'Atendimento a chamados Panco',
                        'Automatização da ordem do frete',
                        'Chamados de sustentação',
                        'Chamados de sustentação',
                        'Automatização da ordem do frete',
                    ][:len(st.session_state.details_df)]
            st.rerun()
    with col_b:
        if st.button('Limpar', use_container_width=True):
            st.session_state.activities_df = blank_activities()
            st.session_state.weekly_df = blank_weekly()
            st.session_state.details_df = blank_details()
            st.rerun()

    st.markdown('---')
    uploaded_state = st.file_uploader('Abrir arquivo salvo (.json)', type=['json'])
    if uploaded_state is not None:
        load_json_payload(uploaded_state.getvalue())
        st.success('Arquivo carregado.')

    st.download_button(
        'Salvar estrutura do apontamento (.json)',
        data=save_json_payload(),
        file_name='ers_timesheet_salvo.json',
        mime='application/json',
        use_container_width=True,
    )

st.info('Fluxo: 1) selecionar atividades, 2) lançar horas da semana, 3) detalhar o apontamento, 4) baixar os relatórios.')

projects = sorted([p for p in normalize_df(st.session_state.weekly_df, WEEKLY_COLUMNS).get('Projeto', pd.Series(dtype=str)).dropna().astype(str).unique() if p])
with st.sidebar:
    st.subheader('Filtro do relatório simples')
    st.session_state.simple_project = st.selectbox('Projeto', [''] + projects, index=([''] + projects).index(st.session_state.simple_project) if st.session_state.simple_project in [''] + projects else 0)
    activity_source = normalize_df(st.session_state.weekly_df, WEEKLY_COLUMNS)
    if st.session_state.simple_project:
        activity_source = activity_source[activity_source['Projeto'] == st.session_state.simple_project]
    activities = sorted([a for a in activity_source.get('Atividade', pd.Series(dtype=str)).dropna().astype(str).unique() if a])
    st.session_state.simple_activity = st.selectbox('Atividade', [''] + activities, index=([''] + activities).index(st.session_state.simple_activity) if st.session_state.simple_activity in [''] + activities else 0)

metrics = weekly_metrics(st.session_state.weekly_df)
m1, m2, m3, m4 = st.columns(4)
m1.metric('Horas na semana', f"{metrics['Total']:.2f}")
m2.metric('Horas segunda', f"{metrics['Seg']:.2f}")
m3.metric('Horas sexta', f"{metrics['Sex']:.2f}")
m4.metric('Atividades selecionadas', int(len(normalize_df(st.session_state.weekly_df, WEEKLY_COLUMNS))))

warnings = validate_weekly(st.session_state.weekly_df)
for msg in warnings:
    st.warning(msg)


tab1, tab2, tab3, tab4 = st.tabs(['1. Seleção de atividades', '2. Lançamento semanal', '3. Detalhe do apontamento', '4. Relatórios'])

with tab1:
    st.markdown('### Janela 1 · Selecionar atividades')
    st.write('Cadastre e marque as atividades que estarão disponíveis na semana.')
    edited_activities = st.data_editor(
        normalize_df(st.session_state.activities_df, ACTIVITY_COLUMNS),
        num_rows='dynamic',
        use_container_width=True,
        key='activities_editor',
        column_config={
            'Selecionar': st.column_config.CheckboxColumn(default=False),
            'Início': st.column_config.DateColumn(format='DD/MM/YYYY'),
            'Fim': st.column_config.DateColumn(format='DD/MM/YYYY'),
        },
    )
    st.session_state.activities_df = edited_activities
    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button('Aplicar seleção na semana', type='primary', use_container_width=True):
            sync_weekly_from_activities()
            st.success('Atividades sincronizadas com a semana.')
            st.rerun()
    with c2:
        st.caption('Dica: sempre clique em “Aplicar seleção na semana” quando alterar a lista de atividades.')

with tab2:
    st.markdown('### Janela 2 · Lançamento semanal')
    week_end = st.session_state.week_start + timedelta(days=6)
    st.write(f"Semana de {st.session_state.week_start.strftime('%d/%m/%Y')} até {week_end.strftime('%d/%m/%Y')}")
    edited_weekly = st.data_editor(
        normalize_df(st.session_state.weekly_df, WEEKLY_COLUMNS),
        num_rows='dynamic',
        use_container_width=True,
        key='weekly_editor',
        column_config={
            'Dom': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
            'Seg': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
            'Ter': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
            'Qua': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
            'Qui': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
            'Sex': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
            'Sáb': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
        },
    )
    st.session_state.weekly_df = edited_weekly
    dcols = st.columns(8)
    for idx, label in enumerate(['Dom', 'Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Total']):
        value = metrics[label] if label != 'Total' else metrics['Total']
        dcols[idx].metric(label, f'{value:.2f}')
    if st.button('Atualizar detalhes automaticamente', use_container_width=False):
        sync_details_from_weekly()
        st.success('Detalhes atualizados a partir do lançamento semanal.')
        st.rerun()

with tab3:
    st.markdown('### Janela 3 · Detalhe do apontamento')
    st.write('Aqui entram os comentários, marcação de restrito e horas reconhecidas.')
    sync_details_from_weekly()
    edited_details = st.data_editor(
        normalize_df(st.session_state.details_df, DETAIL_COLUMNS),
        num_rows='dynamic',
        use_container_width=True,
        key='details_editor',
        column_config={
            'Data': st.column_config.DateColumn(format='DD/MM/YYYY'),
            'Horas Reconhecidas': st.column_config.NumberColumn(min_value=0.0, max_value=24.0, step=0.5, format='%.2f'),
            'Restrito': st.column_config.SelectboxColumn(options=['Não', 'Sim']),
        },
    )
    st.session_state.details_df = edited_details

with tab4:
    st.markdown('### Relatórios gerados')
    report_simple, report_comp = build_reports()
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('**Relatório de horas simples**')
        st.dataframe(report_simple, use_container_width=True, hide_index=True)
    with c2:
        st.markdown('**Relatório de horas compostas**')
        st.dataframe(report_comp, use_container_width=True, hide_index=True)

    simple_bytes = workbook_from_report(report_simple, 'Relatório de Horas Simples', st.session_state.week_start, st.session_state.manager)
    comp_bytes = workbook_from_report(report_comp, 'Relatório de Horas Compostas', st.session_state.week_start, st.session_state.manager)
    package = BytesIO()
    with zipfile.ZipFile(package, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('relatorio_horas_simples.xlsx', simple_bytes.getvalue())
        zf.writestr('relatorio_horas_compostas.xlsx', comp_bytes.getvalue())
    package.seek(0)

    b1, b2, b3 = st.columns(3)
    b1.download_button('Baixar relatório simples', simple_bytes.getvalue(), 'relatorio_horas_simples.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True)
    b2.download_button('Baixar relatório composto', comp_bytes.getvalue(), 'relatorio_horas_compostas.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True)
    b3.download_button('Baixar pacote completo (.zip)', package.getvalue(), 'ers_timesheet_relatorios.zip', 'application/zip', use_container_width=True)
